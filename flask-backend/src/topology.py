# Copyright 2023 Dynatrace LLC

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

#      https://www.apache.org/licenses/LICENSE-2.0

#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.

from datetime import datetime
import json
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import credentials
import dirs
import topology_test
import ui_api
import concurrent.futures

get_z_os_lpar_payload = {
    "allowedToEditFilter": True,
    "listSortType": "NAME",
    "customName": "Hosts",
    "listFiltersPerMEIndex": [
        {
            "indexEntityType": "MONITORED_HOST",
            "filterItems": [
                {
                    "type": "OS_TYPE",
                    "filteredValuesCount": 1,
                    "totalValuesCount": 1,
                    "nameLookupMap": {},
                    "valuesWithDynamicName": {},
                    "values": ["OS_TYPE_ZOS"],
                }
            ],
        }
    ],
    "unsupportedFilters": [],
}


def extract_topology(tenant_key):
    topology = {}

    gen_topology(tenant_key)
    # topology_test.do_test()

    """
    try:
        with open(
            dirs.get_tenant_data_cache_dir() + "/topology.json", "r", encoding="UTF-8"
        ) as f:
            topology = json.load(f)
    except FileNotFoundError as e:
        save_tenant_list(topology)
    """

    return topology


MAX_THREADS = 5
done_count = 0


def gen_topology(tenant_key):
    config = credentials.get_ui_api_call_credentials(tenant_key)
    
    # credentials.set_timeframe(config, 1697104860000, 1697137260000)

    hosts = get_hosts(config)

    pgs = process_hosts(config, hosts)

    services = process_pgs(config, pgs)

    process_backtraces(config, services)

    workbook = openpyxl.Workbook()

    sheet_1 = workbook.active
    sheet_1.title = "app-mf"
    search_for_app_to_mf_func = build_search_for_func(is_application_func)
    excel_page_1 = analyze_and_create_page(
        services, search_for_app_to_mf_func, "APP", "APP-TYPE", "APP_ID"
    )
    save_to_sheet(sheet_1, excel_page_1)

    print(f"keys: {list(services.keys())}")

    sheet_2 = workbook.create_sheet(title="distr-mf")
    is_distr_to_mf_func = build_is_distr_to_mf_func(services)
    search_for_distr_to_mf_func = build_search_for_func(is_distr_to_mf_func)
    excel_page_2 = analyze_and_create_page(
        services, search_for_distr_to_mf_func, "DISTR", "DISTR-TYPE", "DISTR_ID"
    )
    save_to_sheet(sheet_2, excel_page_2)

    save_to_excel(config, workbook)


def analyze_and_create_page(
    services, search_for_func, name_label, type_label, id_label
):
    app_service = analyze_backtraces(services, search_for_func)

    print(app_service)
    excel_page = []

    for found_id, found_list in app_service.items():
        for service_id, service in found_list["services"].items():
            print(
                f'"{found_id}","{found_list["name"]}","{service_id}","{service["name"]}"'
            )
            excel_page.append(
                {
                    name_label: found_list["name"],
                    type_label: found_list["type"],
                    "SERVICE": service["name"],
                    "TYPE": service["type"],
                    id_label: found_id,
                    "SERV_ID": service_id,
                }
            )

    return excel_page


def analyze_backtraces(services, search_for_func):
    app_service = {}

    for service_id, service in services.items():
        if (
            "backtrace" in service
            and "root" in service["backtrace"]
            and "callers" in service["backtrace"]["root"]
        ):
            app_service = search_for_func(
                service_id,
                service["backtrace"]["root"]["serviceName"],
                service["backtrace"]["root"]["serviceType"],
                service["backtrace"]["root"],
                app_service,
            )

    return app_service


def save_to_excel(config, workbook):
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    path = dirs.forward_slash_join(
        dirs.get_tenant_work_tenant_dir(config), f"topology_{current_datetime}.xlsx"
    )

    workbook.save(path)
    print(f"Saved to: {path}")


def save_to_sheet(sheet, data):
    if(len(data) > 0):
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx, value=header)

        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                sheet.cell(row=row_idx, column=col_idx, value=ILLEGAL_CHARACTERS_RE.sub(r'',row[key]))


def is_application_func(node, has_children):
    return "serviceId" in node and node["serviceId"].startswith("APPLICATION-")


def build_is_distr_to_mf_func(mf_services):
    def is_distr_to_mf_func(node, has_children):
        if "serviceId" in node and node["serviceId"] in mf_services:
            return False
        
        if has_children and "serviceType" in node:
            if node["serviceType"] in ["CicsInteraction", "ImsInteraction"]:
                return False

        return True

    return is_distr_to_mf_func


def build_search_for_func(is_found_func):
    def search_for_func(service_id, service_name, service_type, node, app_service):
        
        has_children = False
        if "callers" in node and len(node["callers"]) > 0:
            has_children = True
        
        if is_found_func(node, has_children):
            found_id = node["serviceId"]
            if found_id in app_service:
                pass
            else:
                app_service[found_id] = {
                    "name": node["serviceName"],
                    "type": node["serviceType"],
                    "services": {},
                }

            if service_id in app_service[found_id]["services"]:
                pass
            else:
                app_service[found_id]["services"][service_id] = {
                    "name": service_name,
                    "type": service_type,
                }

        elif has_children:
            for caller in node["callers"]:
                app_service = search_for_func(
                    service_id, service_name, service_type, caller, app_service
                )

        return app_service

    return search_for_func


def process_hosts(config, hosts):
    pgs = {}

    for host in hosts:
        host_PGs = get_host_PGs(config, host["id"])
        for host_pg in host_PGs:
            pg_id = host_pg["id"]
            if pg_id in pgs:
                pass
            else:
                pgs[pg_id] = {"done": False}

    return pgs


def process_pgs(config, pgs):
    services = {}

    for pg_id in pgs.keys():
        pg_service_ids = get_PG_services(config, pg_id)
        for service_id in pg_service_ids:
            if service_id in services:
                pass
            else:
                services[service_id] = {"done": False}
        pgs[pg_id]["done"] = True

    return services


def get_hosts(config):
    response = ui_api.get(
        config,
        ui_api.get_hosts,
        f"?{config['global_timeframe']}&gf=all",
        json.dumps(get_z_os_lpar_payload),
        method="POST",
    )

    zosLpars = response.json()
    print(f"Done: Get LPAR Hosts")

    if "entities" in zosLpars and len(zosLpars["entities"]) > 0:
        return zosLpars["entities"]

    return []


def get_host_PGs(config, host_id):
    response = ui_api.get(
        config,
        ui_api.get_entities,
        f"/{host_id}?fields=%2Bproperties&{config['timeframe_from_to']}",
    )

    host = response.json()
    print(f"Done: {host_id}")

    l1 = "toRelationships"
    l2 = "runsOn"

    if l1 in host and l2 in host[l1] and len(host[l1][l2]) > 0:
        return host[l1][l2]

    return []


def get_PG_services(config, pg_id):
    response = ui_api.get(
        config,
        ui_api.get_pg,
        f"/{pg_id}?{config['timeframe']}&{config['global_timeframe']}",
    )

    pg = response.json()
    print(f"Done: {pg_id}")

    l1 = "services"

    if l1 in pg and len(list(pg[l1].keys())) > 0:
        return list(pg[l1].keys())

    return []


def process_backtraces(config, services):
    done_count = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = [
            executor.submit(get_service_backtrace, config, service_id)
            for service_id in services.keys()
        ]

        for future in concurrent.futures.as_completed(futures):
            try:
                backtrace, service_id = future.result()
                services[service_id]["done"] = True
                services[service_id]["backtrace"] = backtrace
                done_count += 1
                print(
                    f"Done: {service_id} backtrace ({done_count} of {len(list(services.keys()))})"
                )

            except Exception as e:
                print(f"An error occurred: {e}")


def get_service_backtrace(config, service_id):
    response = ui_api.get_wait(
        config,
        ui_api.get_backtrace,
        f"?serviceId={service_id}&sci={service_id}&timeframe={config['timeframe']}",
    )

    backtrace = response.json()

    return backtrace, service_id


def save_tenant_list(payload):
    with open(
        dirs.get_tenant_data_cache_dir() + "/topology.json", "w", encoding="UTF-8"
    ) as f:
        f.write(json.dumps(payload))
